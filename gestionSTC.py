import openpyxl
import os.path
from datetime import date
from openpyxl.reader.excel import load_workbook
from ttkbootstrap import*
import tkinter as tk
from tkinter import messagebox
from ttkbootstrap.tableview import Tableview
from ttkbootstrap.dialogs import Messagebox

class STCInterface(Window):
    def __init__(self):
        super().__init__(size=(1600,1000),title="Manipulation des données STC",minsize=(1000,800),maxsize=(1600,800))
        self.workbook = None
        self.excel_file_path = r"C:\Users\HP\Downloads\stg_SEBN\ProjetStg\données.xlsx"
        self.connectToFile()
        self.widgets()
        self.selectedIndex = -1


    def connectToFile(self):
        if os.path.exists(self.excel_file_path):
            self.workbook = load_workbook(self.excel_file_path)

    def entry_label(self,frame,x,y,text,type,vals):
        label=Label(frame,text=text,bootstyle=(INFO))
        label.place(x=x,y=y)

        if type==3:
            entry = Combobox(frame, values=vals)
            entry.place(x=x + 200, y=y)
        elif type==2:
            entry = DateEntry(frame,bootstyle=(INFO))
            entry.place(x=x + 200, y=y)

        else :
            entry = Entry(frame, bootstyle=(INFO))
            entry.place(x=x + 200, y=y)

        return  entry

    def ajouterMTD(self):
        natureDepart = self.natureDepart.get()
        motif = self.motif.get()
        site = self.site.get()
        affectation = self.affectation.get()
        shift = self.shift.get()
        nomPre = self.nomPre.get()
        matricule = self.matricule.get()
        status = self.status.get()
        contrat = self.contrat.get()
        dept = self.dept.get()
        fonction = self.fonction.get()
        dateEMB = datetime.strptime(self.dateEMB.entry.get(), "%d/%m/%Y")
        dateSTC = datetime.strptime(self.dateSTC.entry.get(), "%d/%m/%Y")


       # print("selected_prestation:",selected_prestation,"date:",month_mapping.get(date.month, "UNKNOWN"),"statut:",status,"matricule:",matricule,"nom:",nom,"prenom:",prenom,"bus:",bus,"dept:",dept,"kst:",kst,"kv:",kv,"site:",site,"station:",station)

        if str(fonction)=="":
            return  Messagebox.show_error(title="Error",message="le champs Fonction est requis",alert=True,padding=(20, 20))
        if str(natureDepart)=="":
            return  Messagebox.show_error(title="Error",message="le champs nature de départ est requis",alert=True,padding=(20, 20))
        if str(motif)=="":
            return  Messagebox.show_error(title="Error",message="le champs motif est requis",alert=True,padding=(20, 20))
        if str(dept)=="":
            return  Messagebox.show_error(title="Error",message="le champs département est requis",alert=True,padding=(20, 20))

        sheet = self.workbook['STC']

        if self.selectedIndex != -1:

            row_to_update = sheet[self.selectedIndex+1]
            row_to_update[1].value = matricule
            row_to_update[2].value = nomPre
            row_to_update[3].value = dateEMB.date()
            row_to_update[4].value = dateSTC.date()
            row_to_update[5].value = natureDepart
            row_to_update[6].value = motif
            row_to_update[7].value = status
            row_to_update[8].value = fonction
            row_to_update[9].value = dept
            row_to_update[10].value = site
            row_to_update[11].value = affectation
            row_to_update[12].value = contrat
            row_to_update[13].value = shift

        else:

            new_row = [(sheet.max_row),matricule, nomPre, dateEMB.date(),dateSTC.date(),natureDepart,motif, status,
                       fonction,dept, site, affectation, contrat,shift]
            # Append the new row to the worksheet
            sheet.append(new_row)

            # Save the workbook to persist the changes
        print(self.workbook.save(self.excel_file_path))  # Specify the file path of the existing Excel file
        self.update_tableview()
        if self.selectedIndex != -1:
            messagebox.showinfo("Success", "Data updated successfully")
        else:
            messagebox.showinfo("Success", "Data inserted successfully")
        self.vider_table()
        self.selectedIndex = -1

    def widgets(self):
        frame = Frame(self)
        frame.pack(side = TOP, fill=BOTH,expand=True)

        frame1 = Frame(frame,bootstyle=INFO)
        frame1.place(x=5,y=0,width=800,height=790)

        lblFrame1 = LabelFrame(frame1,text="manipulation d'un enregistrement",bootstyle=INFO,padding=10)
        lblFrame1.pack(side = TOP, fill=BOTH,expand=True)

        #formulaire

        self.matricule = self.entry_label(lblFrame1, 5, 0, "Matricule", -1, [])
        self.nomPre = self.entry_label(lblFrame1, 5, 50, "Nom & Prenom", -1, [])
        self.contrat = self.entry_label(lblFrame1, 5, 100, "Contrat", 3, ["CDI","CD1","CD2","ANA","INT"])

        self.dept = self.entry_label(lblFrame1, 5, 150, "Département", 3, ["PPE", "PCP"])

        self.fonction = self.entry_label(lblFrame1, 5, 200, "Fonction", -1, [])
        self.status = self.entry_label(lblFrame1, 5, 250, "Statut", 3, ["ADM","IND"])
        self.affectation = self.entry_label(lblFrame1, 5, 300, "Affectation", -1, [])
        self.shift = self.entry_label(lblFrame1, 5, 350, "Shift", 3, ["A","B","C","ABC","ADM"])
        self.dateEMB = self.entry_label(lblFrame1, 5, 400, "Date d'embauche", 2, [])
        self.dateSTC = self.entry_label(lblFrame1, 5, 450, "Date STC", 2, [])
        self.site = self.entry_label(lblFrame1, 5, 500, "Site", 3, ["SEBN-MA1","SEBN-MA2","SETELLITE-1","S-booknadel"])
        self.natureDepart = self.entry_label(lblFrame1, 5, 550, "Nature de départ", -1, [])
        self.motif = self.entry_label(lblFrame1, 5, 600, "Motif de départ", -1, [])

        self.btnSave = Button(lblFrame1, text="Valider", command=self.ajouterMTD, bootstyle="INFO")
        self.btnSave.place(x=250,y=650,width= 215 )

        self.btnUpdate = Button(lblFrame1, text="Modifier", command=self.updateValue, bootstyle="success")
        self.btnUpdate.configure(state="disable")
        self.btnUpdate.place(x=5, y=650, width=215)

        self.btnDrop = Button(lblFrame1, text="Supprimer", command=self.confirm_delete_row, bootstyle="DANGER")
        self.btnDrop.configure(state="disable")
        self.btnDrop.place(x=140, y=700, width=215)


        frame2 = Frame(frame, bootstyle=DANGER)
        frame2.place(x=500, y=0, width=1095, height=790)
        self.lblFrame2 = LabelFrame(frame2, text="Visualization des enregitrements", bootstyle=INFO)
        self.lblFrame2.pack(side=TOP, fill=BOTH, expand=True)

        self.update_tableview()

    def update_tableview(self):
        sheet = self.workbook['STC']

        header_row = [cell.value for cell in sheet[1]]

        data_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data_rows.append(list(row))

        coldata = [{"text": header, "stretch": True} for header in header_row]

        # If data_rows is provided, update the Tableview with new data
        if data_rows is not None:
           # self.tableView.clear_table_data()  # Clear existing data
            if hasattr(self, 'tableView') and self.tableView:
                self.tableView.destroy()
            self.tableView = Tableview(
                self.lblFrame2,
                paginated=True,
                searchable=True,
                bootstyle=(INFO),
                stripecolor=("#F0F0F0", None),
                autoalign=True,
                autofit=True,

                pagesize=23,
                delimiter=";"
            )
            self.tableView.pack(fill=BOTH, expand=True, padx=5, pady=5)
            self.tableView.build_table_data(coldata, data_rows)

            self.tableView.autoalign_columns()
            self.tableView.autofit_columns()
            self.tableView.view.bind("<Button-1>", self.handle_row_click)

    def updateValue(self):
        self.vider_table()
        selected_item_id = self.tableView.view.selection()  # Get the selected item's identifier
        data = self.tableView.view.item(selected_item_id)["values"]
        self.selectedIndex = data[0]

        if data:

            self.dept.insert(0, data[5])
            self.matricule.insert(0, data[1])
            self.nomPre.insert(0, data[2])
            parsed_date1 = datetime.strptime(data[3], "%Y-%m-%d %H:%M:%S")
            parsed_date2 = datetime.strptime(data[4], "%Y-%m-%d %H:%M:%S")

            self.dateEMB.entry.insert(0, parsed_date1.strftime("%d/%m/%Y"))
            self.dateSTC.entry.insert(0, parsed_date2.strftime("%d/%m/%Y"))

            self.natureDepart.insert(0, data[5])
            self.motif.insert(0, data[6])
            self.status.insert(0, data[7])
            self.fonction.insert(0, data[8])
            self.dept.insert(0, data[9])
            self.site.insert(0, data[10])
            self.affectation.insert(0, data[11])
            self.contrat.insert(0, data[12])
            self.shift.insert(0, data[13])



    def handle_row_click(self, event):
        self.btnUpdate.configure(state="enable")
        self.btnDrop.configure(state="enable")

    def confirm_delete_row(self):

        result = messagebox.askquestion("Confirm Deletion", "Are you sure you want to delete this row?")

        if result == "yes":

            self.supprimerLigne()
        else:
            # User canceled, do nothing
            pass

    def supprimerLigne(self):
        selected_item_id = self.tableView.view.selection()  # Get the selected item's identifier
        data = self.tableView.view.item(selected_item_id)["values"]
        self.selectedIndex = data[0]+1
        print(self.selectedIndex)
        sheet = self.workbook['STC']

        if self.selectedIndex != -1:
            sheet.delete_rows(self.selectedIndex)
            self.workbook.save(self.excel_file_path)  # Specify the file path of the existing Excel file
            self.update_tableview()
            messagebox.showinfo("Success", "Data deleted successfully")
            self.selectedIndex = -1

    def vider_table(self):
        self.matricule.delete(0, 'end')
        self.status.delete(0, 'end')
        self.natureDepart.delete(0, 'end')
        self.motif.delete(0, 'end')
        self.fonction.delete(0, 'end')
        self.nomPre.delete(0, 'end')
        self.dept.delete(0, 'end')
        self.site.delete(0, 'end')
        self.affectation.delete(0, 'end')
        self.contrat.delete(0, 'end')
        self.shift.delete(0, 'end')
        self.dateSTC.entry.delete(0, 'end')
        self.dateEMB.entry.delete(0, 'end')

