import openpyxl
import os.path
from datetime import date
from openpyxl.reader.excel import load_workbook
from ttkbootstrap import*
import tkinter as tk
from tkinter import messagebox
from ttkbootstrap.tableview import Tableview
from ttkbootstrap.dialogs import Messagebox

class RestaurationInterface(Window):
    def __init__(self):
        super().__init__(size=(1600,1000),title="Manipulation des données de restauration",minsize=(1000,800),maxsize=(1600,800))
        self.workbook = None
        self.excel_file_path = r"C:\Users\HP\Downloads\stg_SEBN\ProjetStg\data.xlsx"
        self.connectToFile()
        self.widgets()
        self.selectedIndex = -1


    def connectToFile(self):
        if os.path.exists(self.excel_file_path):
            self.workbook = load_workbook(self.excel_file_path)

    def entry_label(self,frame,x,y,text,type,vals):
        label=Label(frame,text=text,bootstyle=(INFO))
        label.place(x=x,y=y)

        if type==3:
            entry = Combobox(frame, values=vals)
            entry.place(x=x + 200, y=y)
        elif type==2:
            entry = DateEntry(frame,bootstyle=(INFO))
            entry.place(x=x + 200, y=y)
        elif type == 4:
            entry = Combobox(frame, values=vals,bootstyle=(INFO))
            entry.place(x=x + 200, y=y)
            entry.bind("<<ComboboxSelected>>", self.update_kst_values)
        elif type == 5:
            entry = Combobox(frame, bootstyle=(INFO))
            entry.place(x=x + 200, y=y)

        else :
            entry = Entry(frame, bootstyle=(INFO))
            entry.place(x=x + 200, y=y)

        return  entry

    def ajouterMTD(self):
        nomPre = self.nomPre.get()
        matricule = self.matricule.get()
        status = self.status.get()
        cout = self.cout.get()
        dept = self.dept.get()
        kst = self.kst.get()
        date = datetime.strptime(self.date.entry.get(), "%d/%m/%Y")

        selected_prestation= self.selected_prestation.get()

       # print("selected_prestation:",selected_prestation,"date:",month_mapping.get(date.month, "UNKNOWN"),"statut:",status,"matricule:",matricule,"nom:",nom,"prenom:",prenom,"bus:",bus,"dept:",dept,"kst:",kst,"kv:",kv,"site:",site,"station:",station)

        if str(status)=="":
            return  Messagebox.show_error(title="Error",message="le champs statut est requis",alert=True,padding=(20, 20))
        if str(dept)=="":
            return  Messagebox.show_error(title="Error",message="le champs département est requis",alert=True,padding=(20, 20))
        if len(cout) == 0 or not cout.isdigit():
            self.cout.delete(0, 'end')
            return Messagebox.show_error(title="Error", message="le champs Cout doit être un entier ", alert=True,
                                  padding=(20, 20))
        sheet = self.workbook['Restauration']

        if self.selectedIndex != -1:

            row_to_update = sheet[self.selectedIndex+1]
            row_to_update[1].value = matricule
            row_to_update[2].value = nomPre
            row_to_update[3].value = date.date()
            row_to_update[4].value = status
            row_to_update[5].value = dept
            row_to_update[6].value = kst
            row_to_update[7].value = cout
            row_to_update[8].value = selected_prestation

        else:

            new_row = [(sheet.max_row),matricule, nomPre, date.date(), status, dept, kst, cout, selected_prestation]
            # Append the new row to the worksheet
            sheet.append(new_row)

            # Save the workbook to persist the changes
        print(self.workbook.save(self.excel_file_path))  # Specify the file path of the existing Excel file
        self.update_tableview()
        if self.selectedIndex != -1:
            messagebox.showinfo("Success", "Data updated successfully")
        else:
            messagebox.showinfo("Success", "Data inserted successfully")
        self.vider_table()
        self.selectedIndex = -1

    def update_kst_values(self, event):
        department = self.dept.get()
        if department == "PPE":
            kst_values = ["123", "124", "125"]
        elif department == "PCP":
            kst_values = ["201", "202", "203"]
        else:
            kst_values = []

        self.kst['values'] = kst_values

    def widgets(self):
        frame = Frame(self)
        frame.pack(side = TOP, fill=BOTH,expand=True)

        frame1 = Frame(frame,bootstyle=INFO)
        frame1.place(x=5,y=0,width=800,height=790)

        lblFrame1 = LabelFrame(frame1,text="manipulation d'un enregistrement",bootstyle=INFO,padding=10)
        lblFrame1.pack(side = TOP, fill=BOTH,expand=True)

        #formulaire

        self.matricule = self.entry_label(lblFrame1, 5, 0, "Matricule",-1,[])
        self.nomPre = self.entry_label(lblFrame1, 5, 70, "Nom & Prénom",-1,[])
        self.cout = self.entry_label(lblFrame1, 5, 140, "Coût",-1,[])

        self.dept = self.entry_label(lblFrame1, 5, 210, "Département", 4, ["PPE","PCP"])

        self.kst = self.entry_label(lblFrame1, 5, 280, "KST", 5, [])
        self.status = self.entry_label(lblFrame1, 5, 350, "Statut",-1,[])

        self.date = self.entry_label(lblFrame1, 5, 420, "Date", 2,[])

        label = Label(frame, text="Prestation", bootstyle=(INFO))
        label.place(x=10, y=550)

        self.selected_prestation = tk.StringVar()
        self.prestation1 = Radiobutton(bootstyle="INFO", text="Sub",variable=self.selected_prestation,value="Subvention")
        self.prestation2 = Radiobutton(bootstyle="INFO", text="HS",variable=self.selected_prestation,value="Heures Supplémentaires")
        self.prestation3 = Radiobutton(bootstyle="INFO", text="Stg",variable=self.selected_prestation,value="Stagiaire")
        self.prestation4 = Radiobutton(bootstyle="INFO", text="PC",variable=self.selected_prestation,value="Pauses Café")

        self.prestation1.place(x=200, y=550)
        self.prestation2.place(x=280, y=550)
        self.prestation3.place(x=360, y=550)
        self.prestation4.place(x=440, y=550)


        self.btnSave = Button(lblFrame1, text="Ajouter", command=self.ajouterMTD, bootstyle="INFO")
        self.btnSave.place(x=250,y=600,width= 215 )

        self.btnUpdate = Button(lblFrame1, text="Modifier", command=self.updateValue, bootstyle="success")
        self.btnUpdate.configure(state="disable")
        self.btnUpdate.place(x=5, y=600, width=215)

        self.btnDrop = Button(lblFrame1, text="Supprimer", command=self.confirm_delete_row, bootstyle="DANGER")
        self.btnDrop.configure(state="disable")
        self.btnDrop.place(x=140, y=680, width=215)
        #btnSave = Button(lblFrame1, text="Valider", command="/")
        #btnSave.place(x=5, y=500, width=200)

        frame2 = Frame(frame, bootstyle=DANGER)
        frame2.place(x=500, y=0, width=1095, height=790)
        self.lblFrame2 = LabelFrame(frame2, text="Visualization des enregitrements", bootstyle=INFO)
        self.lblFrame2.pack(side=TOP, fill=BOTH, expand=True)

        self.update_tableview()

    def update_tableview(self):
        sheet = self.workbook['Restauration']

        header_row = [cell.value for cell in sheet[1]]

        data_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data_rows.append(list(row))

        coldata = [{"text": header, "stretch": True} for header in header_row]

        # If data_rows is provided, update the Tableview with new data
        if data_rows is not None:
           # self.tableView.clear_table_data()  # Clear existing data
            if hasattr(self, 'tableView') and self.tableView:
                self.tableView.destroy()
            self.tableView = Tableview(
                self.lblFrame2,
                paginated=True,
                searchable=True,
                bootstyle=(INFO),
                stripecolor=("#F0F0F0", None),
                autoalign=True,
                autofit=True,

                pagesize=23,
                delimiter=";"
            )
            self.tableView.pack(fill=BOTH, expand=True, padx=5, pady=5)
            self.tableView.build_table_data(coldata, data_rows)

            self.tableView.autoalign_columns()
            self.tableView.autofit_columns()
            self.tableView.view.bind("<Button-1>", self.handle_row_click)

    def updateValue(self):
        self.vider_table()
        selected_item_id = self.tableView.view.selection()  # Get the selected item's identifier
        data = self.tableView.view.item(selected_item_id)["values"]
        self.selectedIndex = data[0]

        if data:

            self.dept.insert(0, data[5])
            self.matricule.insert(0, data[1])
            self.status.insert(0, data[4])
            self.nomPre.insert(0,data[2])
            self.cout.insert(0, data[7])
            self.kst.insert(0, data[6])
            parsed_date = datetime.strptime(data[3], "%Y-%m-%d %H:%M:%S")
            formatted_date = parsed_date.strftime("%d/%m/%Y")
            self.date.entry.insert(0,formatted_date)
            self.selected_prestation.set(data[8])


    def handle_row_click(self, event):
        self.btnUpdate.configure(state="enable")
        self.btnDrop.configure(state="enable")

    def confirm_delete_row(self):

        result = messagebox.askquestion("Confirm Deletion", "Are you sure you want to delete this row?")

        if result == "yes":

            self.supprimerLigne()
        else:
            # User canceled, do nothing
            pass

    def supprimerLigne(self):
        selected_item_id = self.tableView.view.selection()  # Get the selected item's identifier
        data = self.tableView.view.item(selected_item_id)["values"]
        self.selectedIndex = data[0]+1
        print(self.selectedIndex)
        sheet = self.workbook['Restauration']

        if self.selectedIndex != -1:
            sheet.delete_rows(self.selectedIndex)
            self.workbook.save(self.excel_file_path)  # Specify the file path of the existing Excel file
            self.update_tableview()
            messagebox.showinfo("Success", "Data deleted successfully")
            self.selectedIndex = -1

    def vider_table(self):
        self.matricule.delete(0, 'end')
        self.status.delete(0, 'end')
        self.cout.delete(0, 'end')
        self.kst.delete(0, 'end')
        self.date.entry.delete(0, 'end')
        self.nomPre.delete(0, 'end')
        self.dept.delete(0, 'end')
        self.selected_prestation.set("")