
from tkinter import *
from ttkbootstrap import *
from tkinter import messagebox
from ttkbootstrap.tableview import Tableview
from ttkbootstrap.dialogs import Messagebox
import os.path

from openpyxl.reader.excel import load_workbook
class FirstPage:
    def __init__(self, root):
        self.root = root
        self.root.title("First Page")

        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        window_width = screen_width // 2
        window_height = screen_height // 2
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2

        self.root.geometry('5x10')
        self.root.minsize(1000, 800)
        self.root.maxsize(1600, 800)  # Set the maximum window size
        self.root.resizable(False, False);
        self.bg = PhotoImage(file="sebn.png")
        self.bg_image = Label(self.root, image=self.bg).place(x=0, y=0, relwidth=1, relheight=1)

        self.label = tk.Label(root, text="Bienvenue", font=('Microsoft YaHei UI Light', 15, 'bold'),width=30, pady=10)
        self.label.pack(pady=20)

        btn = tk.Button(root, width=29, pady=20, text='Service Transport', bg='#1f72ba', fg='white', border=0,
                     font=('Microsoft YaHei UI Light', 15, 'bold'), cursor='hand2', command=self.open_second_page).place(
            x=100, y=300)
        btn = tk.Button(root, width=29, pady=20, text='Service Restauration', bg='#1f72ba', fg='white', border=0,
                     font=('Microsoft YaHei UI Light', 15, 'bold'), cursor='hand2', command=self.open_third_page).place(
            x=520, y=300)
        btn = tk.Button(root, width=29, pady=20, text='STC', bg='#1f72ba', fg='white', border=0,
                     font=('Microsoft YaHei UI Light', 15, 'bold'), cursor='hand2', command=self.open_fourth_page).place(x=300,
                                                                                                              y=450)

    def open_second_page(self):
        self.root.withdraw()  # Hide the first page
        second_page = tk.Toplevel(self.root)  # Create a new top-level window
        SecondPage(second_page)

    def open_third_page(self):
        self.  root.withdraw()  # Hide the first page
        third_page = tk.Toplevel(self.root)  # Create a new top-level window
        ThirdPage(third_page)

    def open_fourth_page(self):
        self.root.withdraw()  # Hide the first page
        fourth_page = tk.Toplevel(self.root)  # Create a new top-level window
        FourthPage(fourth_page)

class SecondPage:
    def __init__(self, root):
        self.root = root

        self.root.title("Manipulation des données de Transport")
        self.root.minsize(1000, 800)
        #self.root.maxsize(1600, 800)
        #self.root.size=(1600,700)
        self.workbook = None
        self.excel_file_path = r"C:\Users\HP\Downloads\stg_SEBN\ProjetStg\données.xlsx"
        self.connectToFile()
        self.widgets(self.root)
        self.selectedIndex = -1

    def return_to_first_page(self):
        self.root.destroy()  # Close the second page
        first_page.root.deiconify()  # Show the first page

    def connectToFile(self):

        if os.path.exists(self.excel_file_path):
            self.workbook = load_workbook(self.excel_file_path)

    def entry_label(self, frame, x, y, text, type, vals):
        label = Label(frame, text=text, bootstyle=(INFO))
        label.place(x=x, y=y)

        if type == 3:
            entry = Combobox(frame, values=vals)
            entry.place(x=x + 200, y=y)
        elif type == 2:
            entry = DateEntry(frame, bootstyle=(INFO))
            entry.place(x=x + 200, y=y)
        elif type == 4:
            entry = Combobox(frame, values=vals, bootstyle=(INFO))
            entry.place(x=x + 200, y=y)
            entry.bind("<<ComboboxSelected>>", self.update_kst_values)
        elif type == 5:
            entry = Combobox(frame, bootstyle=(INFO))
            entry.place(x=x + 200, y=y)

        else:
            entry = Entry(frame, bootstyle=(INFO))
            entry.place(x=x + 200, y=y)

        return entry

    def ajouterMTD(self):
        month_mapping = {
            1: "JANVIER",
            2: "FÉVRIER",
            3: "MARS",
            4: "AVRIL",
            5: "MAI",
            6: "JUIN",
            7: "JUILLET",
            8: "AOÛT",
            9: "SEPTEMBRE",
            10: "OCTOBRE",
            11: "NOVEMBRE",
            12: "DÉCEMBRE",
        }
        matricule = self.matricule.get()
        status = self.status.get()
        bus = self.bus.get()
        cout = self.cout.get()
        motif = self.motif.get()
        dept = self.dept.get()
        kst = self.kst.get()
        date = datetime.strptime(self.date.entry.get(), "%d/%m/%Y")

        kv = self.kv.get()
        site = self.site.get()
        station = self.station.get()
        selected_prestation = self.selected_prestation.get()

        # print("selected_prestation:",selected_prestation,"date:",month_mapping.get(date.month, "UNKNOWN"),"statut:",status,"matricule:",matricule,"nom:",nom,"prenom:",prenom,"bus:",bus,"dept:",dept,"kst:",kst,"kv:",kv,"site:",site,"station:",station)

        if str(status) == "":
            return Messagebox.show_error(title="Error", message="le champs statut est requis", alert=True,
                                         padding=(20, 20))
        if str(dept) == "":
            return Messagebox.show_error(title="Error", message="le champs département est requis", alert=True,
                                         padding=(20, 20))
        if str(kv) == "":
            return Messagebox.show_error(title="Error", message="le champs kv est requis", alert=True, padding=(20, 20))
        try:
            int(bus)  # Try to convert to an integer
        except ValueError:
            # Show an error message if it's not a valid integer
            self.bus.delete(0, 'end')
            Messagebox.show_error(title="Error", message="le champs N bus doit être un entier", alert=True,
                                  padding=(20, 20))
        if len(cout) == 0 or not cout.isdigit():
            self.cout.delete(0, 'end')
            return Messagebox.show_error(title="Error", message="le champs Cout doit être un entier ", alert=True,
                                         padding=(20, 20))

        sheet = self.workbook['Transport']
        if self.selectedIndex != -1:

            row_to_update = sheet[self.selectedIndex]

            row_to_update[0].value = month_mapping.get(date.month, "UNKNOWN")
            row_to_update[1].value = kv
            row_to_update[2].value = matricule
            row_to_update[3].value = site
            row_to_update[4].value = station
            row_to_update[5].value = int(bus)
            row_to_update[6].value = kst
            row_to_update[7].value = dept
            row_to_update[8].value = status
            row_to_update[9].value = cout
            row_to_update[10].value = date.year
            row_to_update[11].value = date.date()
            row_to_update[12].value = selected_prestation
            row_to_update[13].value = motif
        else:
            # Create a new row with the data
            new_row = [month_mapping.get(date.month, "UNKNOWN"), kv, matricule, site, station, int(bus), kst, dept,
                       status, cout, date.year, date.date(), selected_prestation, motif]
            # Append the new row to the worksheet
            sheet.append(new_row)

        # Save the workbook to persist the changes
        self.workbook.save(self.excel_file_path)  # Specify the file path of the existing Excel file
        self.update_tableview()
        if self.selectedIndex != -1:
            messagebox.showinfo("Success", "Data updated successfully")
        messagebox.showinfo("Success", "Data inserted successfully")
        self.vider_table()
        self.selectedIndex = -1

    def update_kst_values(self, event):
        # You can define the mapping of departments to KST values here
        department = self.dept.get()
        if department == "PPE":
            kst_values = ["41000", "42000", "43000","44000", "45000", "46000","47000", "48000", "49000"]
        elif department == "PCP":
            kst_values = ["13000", "14000", "15000","16000", "17000", "18000","19000", "23000", "24000"]
        elif department == "CPE":
            kst_values = ["23400", "11430", "11400","11600", "11340", "11300","11450", "11304", "11402"]
        elif department == "IT":
            kst_values = ["36530", "35340", "15400","16500", "17350", "14340","13540", "24400", "23600"]
        elif department == "PHR":
            kst_values = ["76300", "57660", "58600","35600", "35500", "88600","87650", "68500", "76600"]
        elif department == "PLM":
            kst_values = ["16540", "65440", "76540","66500", "56760", "56750","76440", "48650", "67550"]
        elif department == "PQM":
            kst_values = ["45676", "56750", "56650","17550", "15670", "18670","33554", "22400", "45400"]
        elif department == "PGM":
            kst_values = ["34300", "57660", "58600", "35600", "35500", "88600", "87650", "68500", "76600"]
        elif department == "PMC":
            kst_values = ["16540", "65440", "76540", "66500", "56760", "56750", "76440", "48650", "67550"]
        elif department == "PTC":
            kst_values = ["45676", "56750", "56650", "17550", "15670", "18670", "33554", "22400", "45400"]
        elif department == "PCP":
            kst_values = ["45676", "56750", "56650", "17550", "15670", "18670", "33554", "22400", "45400"]
        else:
            kst_values = []

        self.kst['values'] = kst_values

    def update_tableview(self):
        sheet = self.workbook['Transport']

        header_row = [cell.value for cell in sheet[1]]

        data_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data_rows.append(list(row))

        coldata = [{"text": header, "stretch": True} for header in header_row]

        # If data_rows is provided, update the Tableview with new data
        if data_rows is not None:
            # self.tableView.clear_table_data()  # Clear existing data
            if hasattr(self, 'tableView') and self.tableView:
                self.tableView.destroy()
            self.tableView = Tableview(
                self.lblFrame2,
                paginated=True,
                searchable=True,
                bootstyle=(INFO),
                stripecolor=("#F0F0F0", None),
                autoalign=True,
                autofit=True,
                height=30,
                pagesize=23,
                delimiter=";"
            )
            self.tableView.place(x=5, y=0, width=590, height=630)
            self.tableView.build_table_data(coldata, data_rows)

            self.tableView.autoalign_columns()
            self.tableView.autofit_columns()
            self.tableView.view.bind("<Button-1>", self.handle_row_click)

    def updateValue(self):
        self.vider_table()
        selected_item_id = self.tableView.view.selection()  # Get the selected item's identifier
        data = self.tableView.view.item(selected_item_id)["values"]
        self.selectedIndex = self.tableView.view.index(selected_item_id) + 2

        if data:
            self.dept.insert(0, data[8])

            self.matricule.insert(0, data[3])

            self.status.insert(0, data[9])
            self.bus.insert(0, data[6])
            self.cout.insert(0, data[10])
            self.motif.insert(0, data[13])

            self.kst.insert(0, data[7])
            self.date.entry.insert(0, data[11])
            self.kv.insert(0, data[2])
            self.site.insert(0, data[4])
            self.station.insert(0, data[5])
            self.selected_prestation.set(data[12])

        print(self.selectedIndex)

    def handle_row_click(self, event):
        self.btnUpdate.configure(state="enable")
        self.btnDrop.configure(state="enable")

    def confirm_delete_row(self):

        result = messagebox.askquestion("Confirm Deletion", "Are you sure you want to delete this row?")

        if result == "yes":
            print(self.selectedIndex)
            self.supprimerLigne()
        else:
            # User canceled, do nothing
            pass

    def supprimerLigne(self):
        selected_item_id = self.tableView.view.selection()  # Get the selected item's identifier
        self.selectedIndex = self.tableView.view.index(selected_item_id) + 2
        sheet = self.workbook['Transport']
        print("koko")
        print(self.selectedIndex)
        if self.selectedIndex != -1:
            sheet.delete_rows(self.selectedIndex)
            self.workbook.save(self.excel_file_path)  # Specify the file path of the existing Excel file
            self.update_tableview()
            messagebox.showinfo("Success", "Data deleted successfully")
            self.selectedIndex = -1

    def vider_table(self):
        self.dept.set("")
        self.kst.set("")
        self.matricule.delete(0, 'end')

        self.status.delete(0, 'end')
        self.bus.delete(0, 'end')
        self.cout.delete(0, 'end')
        self.motif.delete(0, 'end')
        self.dept.delete(0, 'end')
        self.kst.delete(0, 'end')
        self.date.entry.delete(0, 'end')
        self.kv.set("")
        self.site.delete(0, 'end')
        self.station.delete(0, 'end')
        self.selected_prestation.set("")

    def widgets(self,r):
        frame = Frame(r)
        frame.pack(side=TOP, fill=BOTH, expand=True)

        frame1 = Frame(frame, bootstyle=INFO)
        frame1.place(x=5, y=0, width=1000, height=700)

        lblFrame1 = LabelFrame(frame1, text="Manipulation d'un enregistrement", bootstyle=INFO, padding=10)
        lblFrame1.place(x=0, y=0, width=400, height=700)

        # formulaire

        self.matricule = self.entry_label(lblFrame1, 5, 0, "Matricule", -1, [])
        self.cout = self.entry_label(lblFrame1, 5, 40, "Coût", -1, [])
        self.motif = self.entry_label(lblFrame1, 5, 80, "Motif", -1, [])

        self.dept = self.entry_label(lblFrame1, 5, 120, "Département", 4, ["IT","PPE", "CPE","PPR","PHR","PLM", "PQM","PGM","PMC","PTC","PCP"])

        self.kst = self.entry_label(lblFrame1, 5, 160, "KST", 5, [])
        self.status = self.entry_label(lblFrame1, 5, 200, "Statut", 4, ["ADM","IND","DIR"])

        self.date = self.entry_label(lblFrame1, 5, 240, "Date", 2, [])
        self.kv = self.entry_label(lblFrame1, 5, 280, "KW", 3, [f"KW{i:02}" for i in range(1, 53)])
        self.site = self.entry_label(lblFrame1, 5, 320, "Site", -1, [])
        self.station = self.entry_label(lblFrame1, 5, 360, "Station", -1, [])
        self.bus = self.entry_label(lblFrame1, 5, 400, "N Bus", -1, [])

        label = Label(frame, text="Prestation", bootstyle=(INFO))
        label.place(x=10, y=490)
        self.selected_prestation = tk.StringVar()
        self.prestation1 = Radiobutton(lblFrame1,bootstyle="INFO", text="VN", variable=self.selected_prestation,
                                       value="Voyages Normaux")
        self.prestation2 = Radiobutton(lblFrame1,bootstyle="INFO", text="HS", variable=self.selected_prestation,
                                       value="Heures Supplémentaires")
        self.prestation3 = Radiobutton(lblFrame1,bootstyle="INFO", text="TAXI", variable=self.selected_prestation, value="TAXI")
        self.prestation1.place(x=200, y=460)
        self.prestation2.place(x=250, y=460)
        self.prestation3.place(x=300, y=460)

        self.btnSave = Button(lblFrame1, text="Valider", command=self.ajouterMTD, bootstyle="INFO")
        self.btnSave.place(x=190, y=500, width=150)

        self.btnUpdate = Button(lblFrame1, text="Modifier", command=self.updateValue, bootstyle="success")
        self.btnUpdate.configure(state="disable")
        self.btnUpdate.place(x=5, y=500, width=150)

        self.btnDrop = Button(lblFrame1, text="Supprimer", command=self.confirm_delete_row, bootstyle="DANGER")
        self.btnDrop.configure(state="disable")
        self.btnDrop.place(x=190, y=550, width=150)

        # Create a button with the image
        self.return_button = Button(lblFrame1, text="Page d'Accueil", width=19, bootstyle=SECONDARY, command=self.return_to_first_page)
        self.return_button.place(x=5, y=550, width=150)
        """self.icon_button = Button(lblFrame1, text="Page d'Accueil", width=19, bootstyle=SECONDARY
                                  )
        self.icon_button.place(x=5, y=680)"""

        frame2 = Frame(frame, bootstyle=DANGER)
        frame2.place(x=400, y=0, width=1000, height=700)
        self.lblFrame2 = LabelFrame(frame2, text="Visualization des enregitrements", bootstyle=INFO)
        self.lblFrame2.place(x=0, y=0, width=600, height=700)

        self.update_tableview()

class ThirdPage:
    def __init__(self, root):
        self.root = root

        self.root.title("Manipulation des données de Restauration")
        self.root.minsize(1000, 800)
        #self.root.maxsize(1600, 800)
        #self.root.size=(1600,700)
        self.workbook = None
        self.excel_file_path = r"C:\Users\HP\Downloads\stg_SEBN\ProjetStg\données.xlsx"
        self.connectToFile()
        self.widgets(self.root)
        self.selectedIndex = -1

    def return_to_first_page(self):
        self.root.destroy()  # Close the second page
        first_page.root.deiconify()  # Show the first page

    def connectToFile(self):

        if os.path.exists(self.excel_file_path):
            self.workbook = load_workbook(self.excel_file_path)

    def entry_label(self, frame, x, y, text, type, vals):
        label = Label(frame, text=text, bootstyle=(INFO))
        label.place(x=x, y=y)

        if type == 3:
            entry = Combobox(frame, values=vals)
            entry.place(x=x + 200, y=y)
        elif type == 2:
            entry = DateEntry(frame, bootstyle=(INFO))
            entry.place(x=x + 200, y=y)
        elif type == 4:
            entry = Combobox(frame, values=vals, bootstyle=(INFO))
            entry.place(x=x + 200, y=y)
            entry.bind("<<ComboboxSelected>>", self.update_kst_values)
        elif type == 5:
            entry = Combobox(frame, bootstyle=(INFO))
            entry.place(x=x + 200, y=y)

        else:
            entry = Entry(frame, bootstyle=(INFO))
            entry.place(x=x + 200, y=y)

        return entry

    def ajouterMTD(self):
        nomPre = self.nomPre.get()
        matricule = self.matricule.get()
        status = self.status.get()
        cout = self.cout.get()
        dept = self.dept.get()
        kst = self.kst.get()
        date = datetime.strptime(self.date.entry.get(), "%d/%m/%Y")

        selected_prestation = self.selected_prestation.get()

        # print("selected_prestation:",selected_prestation,"date:",month_mapping.get(date.month, "UNKNOWN"),"statut:",status,"matricule:",matricule,"nom:",nom,"prenom:",prenom,"bus:",bus,"dept:",dept,"kst:",kst,"kv:",kv,"site:",site,"station:",station)

        if str(status) == "":
            return Messagebox.show_error(title="Error", message="le champs statut est requis", alert=True,
                                         padding=(20, 20))
        if str(dept) == "":
            return Messagebox.show_error(title="Error", message="le champs département est requis", alert=True,
                                         padding=(20, 20))
        if len(cout) == 0 or not cout.isdigit():
            self.cout.delete(0, 'end')
            return Messagebox.show_error(title="Error", message="le champs Cout doit être un entier ", alert=True,
                                         padding=(20, 20))
        sheet = self.workbook['Restauration']

        if self.selectedIndex != -1:

            row_to_update = sheet[self.selectedIndex + 1]
            row_to_update[1].value = matricule
            row_to_update[2].value = nomPre
            row_to_update[3].value = date.date()
            row_to_update[4].value = status
            row_to_update[5].value = dept
            row_to_update[6].value = kst
            row_to_update[7].value = cout
            row_to_update[8].value = selected_prestation

        else:

            new_row = [(sheet.max_row), matricule, nomPre, date.date(), status, dept, kst, cout, selected_prestation]
            # Append the new row to the worksheet
            sheet.append(new_row)

            # Save the workbook to persist the changes
        print(self.workbook.save(self.excel_file_path))  # Specify the file path of the existing Excel file
        self.update_tableview()
        if self.selectedIndex != -1:
            messagebox.showinfo("Success", "Data updated successfully")
        else:
            messagebox.showinfo("Success", "Data inserted successfully")
        self.vider_table()
        self.selectedIndex = -1

    def update_kst_values(self, event):
        department = self.dept.get()
        if department == "PPE":
            kst_values = ["123", "124", "125"]
        elif department == "PCP":
            kst_values = ["201", "202", "203"]
        else:
            kst_values = []

        self.kst['values'] = kst_values

    def widgets(self,r):
        frame = Frame(r)
        frame.pack(side=TOP, fill=BOTH, expand=True)

        frame1 = Frame(frame, bootstyle=INFO)
        frame1.place(x=5, y=0, width=1000, height=700)

        lblFrame1 = LabelFrame(frame1, text="manipulation d'un enregistrement", bootstyle=INFO, padding=10)
        lblFrame1.place(x=0, y=0, width=400, height=700)

        # formulaire

        self.matricule = self.entry_label(lblFrame1, 5, 0, "Matricule", -1, [])
        self.nomPre = self.entry_label(lblFrame1, 5, 40, "Nom & Prénom", -1, [])
        self.cout = self.entry_label(lblFrame1, 5, 80, "Coût", -1, [])

        self.dept = self.entry_label(lblFrame1, 5, 120, "Département", 4, ["PPE", "PCP"])

        self.kst = self.entry_label(lblFrame1, 5, 160, "KST", 5, [])
        self.status = self.entry_label(lblFrame1, 5, 200, "Statut", -1, [])

        self.date = self.entry_label(lblFrame1, 5, 240, "Date", 2, [])

        label = Label(frame, text="Prestation", bootstyle=(INFO))
        label.place(x=10, y=300)

        self.selected_prestation = tk.StringVar()
        self.prestation1 = Radiobutton(lblFrame1,bootstyle="INFO", text="Sub", variable=self.selected_prestation,
                                       value="Subvention")
        self.prestation2 = Radiobutton(lblFrame1,bootstyle="INFO", text="HS", variable=self.selected_prestation,
                                       value="Heures Supplémentaires")
        self.prestation3 = Radiobutton(lblFrame1,bootstyle="INFO", text="Stg", variable=self.selected_prestation,
                                       value="Stagiaire")
        self.prestation4 = Radiobutton(lblFrame1,bootstyle="INFO", text="PC", variable=self.selected_prestation,
                                       value="Pauses Café")

        self.prestation1.place(x=200, y=280)
        self.prestation2.place(x=250, y=280)
        self.prestation3.place(x=300, y=280)
        self.prestation4.place(x=350, y=280)

        self.btnSave = Button(lblFrame1, text="Valider", command=self.ajouterMTD, bootstyle="INFO")
        self.btnSave.place(x=190, y=400, width=150)

        self.btnUpdate = Button(lblFrame1, text="Modifier", command=self.updateValue, bootstyle="success")
        self.btnUpdate.configure(state="disable")
        self.btnUpdate.place(x=5, y=400, width=150)

        self.btnDrop = Button(lblFrame1, text="Supprimer", command=self.confirm_delete_row, bootstyle="DANGER")
        self.btnDrop.configure(state="disable")
        self.btnDrop.place(x=190, y=440, width=150)

        self.return_button = Button(lblFrame1, text="Page d'Accueil", width=19, bootstyle=SECONDARY,
                                    command=self.return_to_first_page)
        self.return_button.place(x=5, y=440, width=150)
        # btnSave = Button(lblFrame1, text="Valider", command="/")
        # btnSave.place(x=5, y=500, width=200)

        frame2 = Frame(frame, bootstyle=DANGER)
        frame2.place(x=400, y=0, width=1000, height=700)
        self.lblFrame2 = LabelFrame(frame2, text="Visualization des enregitrements", bootstyle=INFO)
        self.lblFrame2.place(x=0, y=0, width=600, height=700)

        self.update_tableview()

    def update_tableview(self):
        sheet = self.workbook['Restauration']

        header_row = [cell.value for cell in sheet[1]]

        data_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data_rows.append(list(row))

        coldata = [{"text": header, "stretch": True} for header in header_row]

        # If data_rows is provided, update the Tableview with new data
        if data_rows is not None:
            # self.tableView.clear_table_data()  # Clear existing data
            if hasattr(self, 'tableView') and self.tableView:
                self.tableView.destroy()
            self.tableView = Tableview(
                self.lblFrame2,
                paginated=True,
                searchable=True,
                bootstyle=(INFO),
                stripecolor=("#F0F0F0", None),
                autoalign=True,
                autofit=True,

                pagesize=23,
                delimiter=";"
            )
            self.tableView.place(x=5, y=0, width=590, height=630)
            self.tableView.build_table_data(coldata, data_rows)

            self.tableView.autoalign_columns()
            self.tableView.autofit_columns()
            self.tableView.view.bind("<Button-1>", self.handle_row_click)

    def updateValue(self):
        self.vider_table()
        selected_item_id = self.tableView.view.selection()  # Get the selected item's identifier
        data = self.tableView.view.item(selected_item_id)["values"]
        self.selectedIndex = data[0]

        if data:
            self.dept.insert(0, data[5])
            self.matricule.insert(0, data[1])
            self.status.insert(0, data[4])
            self.nomPre.insert(0, data[2])
            self.cout.insert(0, data[7])
            self.kst.insert(0, data[6])
            parsed_date = datetime.strptime(data[3], "%Y-%m-%d %H:%M:%S")
            formatted_date = parsed_date.strftime("%d/%m/%Y")
            self.date.entry.insert(0, formatted_date)
            self.selected_prestation.set(data[8])

    def handle_row_click(self, event):
        self.btnUpdate.configure(state="enable")
        self.btnDrop.configure(state="enable")

    def confirm_delete_row(self):

        result = messagebox.askquestion("Confirm Deletion", "Are you sure you want to delete this row?")

        if result == "yes":

            self.supprimerLigne()
        else:
            # User canceled, do nothing
            pass

    def supprimerLigne(self):
        selected_item_id = self.tableView.view.selection()  # Get the selected item's identifier
        data = self.tableView.view.item(selected_item_id)["values"]
        self.selectedIndex = data[0] + 1
        print(self.selectedIndex)
        sheet = self.workbook['Restauration']

        if self.selectedIndex != -1:
            sheet.delete_rows(self.selectedIndex)
            self.workbook.save(self.excel_file_path)  # Specify the file path of the existing Excel file
            self.update_tableview()
            messagebox.showinfo("Success", "Data deleted successfully")
            self.selectedIndex = -1

    def vider_table(self):
        self.matricule.delete(0, 'end')
        self.status.delete(0, 'end')
        self.cout.delete(0, 'end')
        self.kst.delete(0, 'end')
        self.date.entry.delete(0, 'end')
        self.nomPre.delete(0, 'end')
        self.dept.delete(0, 'end')
        self.selected_prestation.set("")

class FourthPage:
    def __init__(self, root):
        self.root = root

        self.root.title("Manipulation des données d'STC")
        self.root.minsize(1000, 800)
        #self.root.maxsize(1600, 800)
        #self.root.size=(1600,700)
        self.workbook = None
        self.excel_file_path = r"C:\Users\HP\Downloads\stg_SEBN\ProjetStg\données.xlsx"
        self.connectToFile()
        self.widgets(self.root)
        self.selectedIndex = -1

    def return_to_first_page(self):
        self.root.destroy()  # Close the second page
        first_page.root.deiconify()  # Show the first page
    def connectToFile(self):
        if os.path.exists(self.excel_file_path):
            self.workbook = load_workbook(self.excel_file_path)

    def entry_label(self,frame,x,y,text,type,vals):
        label=Label(frame,text=text,bootstyle=(INFO))
        label.place(x=x,y=y)

        if type==3:
            entry = Combobox(frame, values=vals)
            entry.place(x=x + 200, y=y)
        elif type==2:
            entry = DateEntry(frame,bootstyle=(INFO))
            entry.place(x=x + 200, y=y)

        else :
            entry = Entry(frame, bootstyle=(INFO))
            entry.place(x=x + 200, y=y)

        return  entry

    def ajouterMTD(self):
        natureDepart = self.natureDepart.get()
        motif = self.motif.get()
        site = self.site.get()
        affectation = self.affectation.get()
        shift = self.shift.get()
        nomPre = self.nomPre.get()
        matricule = self.matricule.get()
        status = self.status.get()
        contrat = self.contrat.get()
        dept = self.dept.get()
        fonction = self.fonction.get()
        dateEMB = datetime.strptime(self.dateEMB.entry.get(), "%d/%m/%Y")
        dateSTC = datetime.strptime(self.dateSTC.entry.get(), "%d/%m/%Y")


       # print("selected_prestation:",selected_prestation,"date:",month_mapping.get(date.month, "UNKNOWN"),"statut:",status,"matricule:",matricule,"nom:",nom,"prenom:",prenom,"bus:",bus,"dept:",dept,"kst:",kst,"kv:",kv,"site:",site,"station:",station)

        if str(fonction)=="":
            return  Messagebox.show_error(title="Error",message="le champs Fonction est requis",alert=True,padding=(20, 20))
        if str(natureDepart)=="":
            return  Messagebox.show_error(title="Error",message="le champs nature de départ est requis",alert=True,padding=(20, 20))
        if str(motif)=="":
            return  Messagebox.show_error(title="Error",message="le champs motif est requis",alert=True,padding=(20, 20))
        if str(dept)=="":
            return  Messagebox.show_error(title="Error",message="le champs département est requis",alert=True,padding=(20, 20))

        sheet = self.workbook['STC']

        if self.selectedIndex != -1:

            row_to_update = sheet[self.selectedIndex+1]
            row_to_update[1].value = matricule
            row_to_update[2].value = nomPre
            row_to_update[3].value = dateEMB.date()
            row_to_update[4].value = dateSTC.date()
            row_to_update[5].value = natureDepart
            row_to_update[6].value = motif
            row_to_update[7].value = status
            row_to_update[8].value = fonction
            row_to_update[9].value = dept
            row_to_update[10].value = site
            row_to_update[11].value = affectation
            row_to_update[12].value = contrat
            row_to_update[13].value = shift

        else:

            new_row = [(sheet.max_row),matricule, nomPre, dateEMB.date(),dateSTC.date(),natureDepart,motif, status,
                       fonction,dept, site, affectation, contrat,shift]
            # Append the new row to the worksheet
            sheet.append(new_row)

            # Save the workbook to persist the changes
        print(self.workbook.save(self.excel_file_path))  # Specify the file path of the existing Excel file
        self.update_tableview()
        if self.selectedIndex != -1:
            messagebox.showinfo("Success", "Data updated successfully")
        else:
            messagebox.showinfo("Success", "Data inserted successfully")
        self.vider_table()
        self.selectedIndex = -1

    def widgets(self,r):
        frame = Frame(r)
        frame.pack(side = TOP, fill=BOTH,expand=True)

        frame1 = Frame(frame,bootstyle=INFO)
        frame1.place(x=5, y=0, width=1000, height=700)

        lblFrame1 = LabelFrame(frame1,text="manipulation d'un enregistrement",bootstyle=INFO,padding=10)
        lblFrame1.place(x=0, y=0, width=400, height=700)

        #formulaire

        self.matricule = self.entry_label(lblFrame1, 5, 0, "Matricule", -1, [])
        self.nomPre = self.entry_label(lblFrame1, 5, 40, "Nom & Prenom", -1, [])
        self.contrat = self.entry_label(lblFrame1, 5, 80, "Contrat", 3, ["CDI","CD1","CD2","ANA","INT"])

        self.dept = self.entry_label(lblFrame1, 5, 120, "Département", 3, ["PPE", "PCP"])

        self.fonction = self.entry_label(lblFrame1, 5, 160, "Fonction", -1, [])
        self.status = self.entry_label(lblFrame1, 5, 200, "Statut", 3, ["ADM","IND"])
        self.affectation = self.entry_label(lblFrame1, 5, 240, "Affectation", -1, [])
        self.shift = self.entry_label(lblFrame1, 5, 280, "Shift", 3, ["A","B","C","ABC","ADM"])
        self.dateEMB = self.entry_label(lblFrame1, 5, 320, "Date d'embauche", 2, [])
        self.dateSTC = self.entry_label(lblFrame1, 5, 360, "Date STC", 2, [])
        self.site = self.entry_label(lblFrame1, 5, 400, "Site", 3, ["SEBN-MA1","SEBN-MA2","SETELLITE-1","S-booknadel"])
        self.natureDepart = self.entry_label(lblFrame1, 5, 440, "Nature de départ", -1, [])
        self.motif = self.entry_label(lblFrame1, 5, 480, "Motif de départ", -1, [])

        self.btnSave = Button(lblFrame1, text="Valider", command=self.ajouterMTD, bootstyle="INFO")
        self.btnSave.place(x=190,y=520,width= 150 )

        self.btnUpdate = Button(lblFrame1, text="Modifier", command=self.updateValue, bootstyle="success")
        self.btnUpdate.configure(state="disable")
        self.btnUpdate.place(x=5, y=520, width=150)

        self.btnDrop = Button(lblFrame1, text="Supprimer", command=self.confirm_delete_row, bootstyle="DANGER")
        self.btnDrop.configure(state="disable")
        self.btnDrop.place(x=190, y=560, width=150)

        self.return_button = Button(lblFrame1, text="Page d'Accueil", width=19, bootstyle=SECONDARY,
                                    command=self.return_to_first_page)
        self.return_button.place(x=5, y=560, width=150)

        frame2 = Frame(frame, bootstyle=DANGER)
        frame2.place(x=400, y=0, width=1000, height=700)
        self.lblFrame2 = LabelFrame(frame2, text="Visualization des enregitrements", bootstyle=INFO)
        self.lblFrame2.place(x=0, y=0, width=600, height=700)

        self.update_tableview()

    def update_tableview(self):
        sheet = self.workbook['STC']

        header_row = [cell.value for cell in sheet[1]]

        data_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data_rows.append(list(row))

        coldata = [{"text": header, "stretch": True} for header in header_row]

        # If data_rows is provided, update the Tableview with new data
        if data_rows is not None:
           # self.tableView.clear_table_data()  # Clear existing data
            if hasattr(self, 'tableView') and self.tableView:
                self.tableView.destroy()
            self.tableView = Tableview(
                self.lblFrame2,
                paginated=True,
                searchable=True,
                bootstyle=(INFO),
                stripecolor=("#F0F0F0", None),
                autoalign=True,
                autofit=True,

                pagesize=23,
                delimiter=";"
            )
            self.tableView.place(x=5, y=0, width=590, height=630)
            self.tableView.build_table_data(coldata, data_rows)

            self.tableView.autoalign_columns()
            self.tableView.autofit_columns()
            self.tableView.view.bind("<Button-1>", self.handle_row_click)

    def updateValue(self):
        self.vider_table()
        selected_item_id = self.tableView.view.selection()  # Get the selected item's identifier
        data = self.tableView.view.item(selected_item_id)["values"]
        self.selectedIndex = data[0]

        if data:

            self.dept.insert(0, data[5])
            self.matricule.insert(0, data[1])
            self.nomPre.insert(0, data[2])
            parsed_date1 = datetime.strptime(data[3], "%Y-%m-%d %H:%M:%S")
            parsed_date2 = datetime.strptime(data[4], "%Y-%m-%d %H:%M:%S")

            self.dateEMB.entry.insert(0, parsed_date1.strftime("%d/%m/%Y"))
            self.dateSTC.entry.insert(0, parsed_date2.strftime("%d/%m/%Y"))

            self.natureDepart.insert(0, data[5])
            self.motif.insert(0, data[6])
            self.status.insert(0, data[7])
            self.fonction.insert(0, data[8])
            self.dept.insert(0, data[9])
            self.site.insert(0, data[10])
            self.affectation.insert(0, data[11])
            self.contrat.insert(0, data[12])
            self.shift.insert(0, data[13])



    def handle_row_click(self, event):
        self.btnUpdate.configure(state="enable")
        self.btnDrop.configure(state="enable")

    def confirm_delete_row(self):

        result = messagebox.askquestion("Confirm Deletion", "Are you sure you want to delete this row?")

        if result == "yes":

            self.supprimerLigne()
        else:
            # User canceled, do nothing
            pass

    def supprimerLigne(self):
        selected_item_id = self.tableView.view.selection()  # Get the selected item's identifier
        data = self.tableView.view.item(selected_item_id)["values"]
        self.selectedIndex = data[0]

        sheet = self.workbook['STC']

        if self.selectedIndex != -1:
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == data[0]:
                    sheet.delete_rows(row[0].row)

                    self.workbook.save(self.excel_file_path)  # Specify the file path of the existing Excel file
                    self.update_tableview()
                    messagebox.showinfo("Success", "Data deleted successfully")
                    self.selectedIndex = -1

    def vider_table(self):
        self.matricule.delete(0, 'end')
        self.status.delete(0, 'end')
        self.natureDepart.delete(0, 'end')
        self.motif.delete(0, 'end')
        self.fonction.delete(0, 'end')
        self.nomPre.delete(0, 'end')
        self.dept.delete(0, 'end')
        self.site.delete(0, 'end')
        self.affectation.delete(0, 'end')
        self.contrat.delete(0, 'end')
        self.shift.delete(0, 'end')
        self.dateSTC.entry.delete(0, 'end')
        self.dateEMB.entry.delete(0, 'end')


root = tk.Tk()
first_page = FirstPage(root)
root.mainloop()
